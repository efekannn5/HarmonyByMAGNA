"""
Excel Export Service for Dolly Tasks
Kurumsal renklerle profesyonel Excel raporları oluşturur
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage


class ExcelExportService:
    """Excel export service for dolly tasks"""
    
    # Kurumsal renkler (Magna renk paleti - soft & professional)
    COLOR_HEADER = "1E3A8A"        # Koyu mavi (header)
    COLOR_SUBHEADER = "3B82F6"     # Orta mavi (alt başlık)
    COLOR_ROW_EVEN = "EFF6FF"      # Açık mavi (çift satırlar)
    COLOR_ROW_ODD = "FFFFFF"       # Beyaz (tek satırlar)
    COLOR_TEXT_DARK = "1F2937"     # Koyu gri (metin)
    COLOR_BORDER = "D1D5DB"        # Açık gri (kenarlık)
    COLOR_VIN_BG = "FAFAFA"        # VIN detay arka plan
    COLOR_VIN_HEADER = "E5E7EB"    # VIN başlık
    
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin', color=self.COLOR_BORDER),
            right=Side(style='thin', color=self.COLOR_BORDER),
            top=Side(style='thin', color=self.COLOR_BORDER),
            bottom=Side(style='thin', color=self.COLOR_BORDER)
        )
    
    def export_dolly_task(self, part_number: str, submissions: List[Any]) -> io.BytesIO:
        """
        Export dolly task to Excel
        
        Args:
            part_number: Part number (PartNumber)
            submissions: List of DollySubmissionHold entries
            
        Returns:
            BytesIO: Excel file in memory
        """
        if not submissions:
            raise ValueError("No submissions found")
        
        # Excel workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Dolly Listesi"
        
        # Başlık bilgileri ekle
        self._add_header(ws, part_number, submissions)
        
        # Dolly verilerini grupla
        dolly_groups = self._group_by_dolly(submissions)
        
        # Tablo başlıkları ekle (4. satır)
        self._add_table_headers(ws)
        
        # Dolly verilerini ekle
        self._add_dolly_data(ws, dolly_groups)
        
        # Sütun genişliklerini ayarla
        self._adjust_column_widths(ws)
        
        # YENİ: Dolly Order No'ya göre VIN listesi sayfası ekle
        self._add_vin_by_order_sheet(wb, dolly_groups)
        
        # Excel dosyasını memory'de oluştur
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _add_header(self, ws, part_number: str, submissions: List[Any]):
        """Başlık bilgilerini ekle (1-4. satırlar)"""
        
        # Ana başlık (1. satır)
        ws.merge_cells('A1:F1')  # G sütununu logo için ayırdık
        header_cell = ws['A1']
        header_cell.value = f"DOLLY GÖREV DETAYI - {part_number}"
        header_cell.font = Font(size=16, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 🔧 HÜCRE BOYUTU AYARLARI (Logo için)
        ws.row_dimensions[1].height = 55        # ← Satır yüksekliği (height)
        ws.column_dimensions['G'].width = 12    # ← Sütun genişliği (width)
        
        # G1 için arka plan rengi ve hücre ortalaması
        ws['G1'].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        ws['G1'].alignment = Alignment(horizontal="center", vertical="center")  # Hücre içi ortalama
        
        # Magna Logo ekle (sağ üst köşe) - BAŞLIKTAN SONRA
        try:
            # Dinamik logo yolu - orjlogo.png kullan
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            logo_path = os.path.join(project_root, 'app', 'static', 'img', 'orjlogo.png')
            
            if os.path.exists(logo_path):
                img = ExcelImage(logo_path)
                img.height = 75
                img.width = 75
                # G1 hücresine yerleştir
                ws.add_image(img, 'G1')
        except Exception as e:
            # Logo eklenemezse sessizce devam et (Excel yine oluşur)
            pass
        
        # Alt başlık bilgileri (2. satır)
        first_sub = submissions[0]
        
        ws.merge_cells('A2:B2')
        ws['A2'] = "Oluşturulma Tarihi:"
        ws['C2'] = first_sub.CreatedAt.strftime('%d.%m.%Y %H:%M') if hasattr(first_sub, 'CreatedAt') and first_sub.CreatedAt else "-"
        
        ws.merge_cells('D2:E2')
        ws['D2'] = "Durum:"
        ws['F2'] = (first_sub.Status or "PENDING").upper()
        
        ws['G2'] = f"Toplam: {len(submissions)} VIN"
        
        # Alt başlık stilleri
        for cell_ref in ['A2', 'C2', 'D2', 'F2', 'G2']:
            cell = ws[cell_ref]
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER, end_color=self.COLOR_SUBHEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        ws.row_dimensions[2].height = 25
        
        # Sevkiyat bilgileri (3. satır) - Sadece girilmişse göster
        info_parts = []
        if hasattr(first_sub, 'SeferNo') and first_sub.SeferNo:
            info_parts.append(f"Sefer: {first_sub.SeferNo}")
        if hasattr(first_sub, 'PlakaNo') and first_sub.PlakaNo:
            info_parts.append(f"Plaka: {first_sub.PlakaNo}")
        if hasattr(first_sub, 'IrsaliyeNo') and first_sub.IrsaliyeNo:
            info_parts.append(f"İrsaliye: {first_sub.IrsaliyeNo}")
        if hasattr(first_sub, 'Lokasyon') and first_sub.Lokasyon:
            info_parts.append(f"Lokasyon: {first_sub.Lokasyon}")
        
        if info_parts:
            ws.merge_cells('A3:G3')
            info_cell = ws['A3']
            info_cell.value = "   |   ".join(info_parts)
            info_cell.font = Font(size=9, bold=True, color="1F2937")
            info_cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            info_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[3].height = 22
        else:
            # Bilgi yoksa boş satır
            ws.row_dimensions[3].height = 5
        
        # Boş satır
        ws.row_dimensions[4].height = 5
    
    def _add_table_headers(self, ws):
        """Tablo başlıklarını ekle (5. satır)"""
        headers = ["#", "DOLLY ORDER NO", "VIN SAYISI", "DURUM", "EOL", "CUSTOMER", "VIN KIRILIMI"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
        
        ws.row_dimensions[5].height = 25
    
    def _group_by_dolly(self, submissions: List[Any]) -> Dict[str, List[Any]]:
        """Submissions'ları dolly no'ya göre grupla"""
        dolly_groups = defaultdict(list)
        
        for sub in submissions:
            dolly_groups[sub.DollyNo].append(sub)
        
        return dolly_groups
    
    def _add_dolly_data(self, ws, dolly_groups: Dict[str, List[Any]]):
        """Dolly verilerini ekle"""
        row_num = 6  # Başlık satırı 5'e kaydı, veri 6'dan başlasın
        dolly_index = 1
        
        for dolly_no, vin_entries in dolly_groups.items():
            # Renk belirleme (zebra pattern)
            row_color = self.COLOR_ROW_EVEN if dolly_index % 2 == 0 else self.COLOR_ROW_ODD
            
            # İlk VIN entry'den bilgileri al
            first_entry = vin_entries[0]
            
            # Dolly ana satırı
            ws.cell(row=row_num, column=1, value=dolly_index)
            ws.cell(row=row_num, column=2, value=f"{first_entry.DollyOrderNo or dolly_no} (D:{dolly_no})")
            ws.cell(row=row_num, column=3, value=f"{len(vin_entries)} VIN")
            ws.cell(row=row_num, column=4, value=(first_entry.Status or "PENDING").upper())
            ws.cell(row=row_num, column=5, value=first_entry.EOLName or "-")
            ws.cell(row=row_num, column=6, value=first_entry.CustomerReferans or "-")
            ws.cell(row=row_num, column=7, value="Aşağıda detay")
            
            # Ana satır stilleri
            for col in range(1, 8):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal="center" if col in [1, 3, 4] else "left", vertical="center")
                if col == 2:
                    cell.font = Font(bold=True, size=10)
            
            row_num += 1
            
            # VIN Kırılımı başlığı
            ws.merge_cells(f'A{row_num}:G{row_num}')
            vin_header = ws.cell(row=row_num, column=1)
            vin_header.value = f"   ↳ VIN Kırılımı ({len(vin_entries)} adet)"
            vin_header.font = Font(italic=True, bold=True, size=9, color="4B5563")
            vin_header.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            vin_header.alignment = Alignment(horizontal="left", vertical="center")
            vin_header.border = self.thin_border
            row_num += 1
            
            # VIN detay başlıkları
            vin_headers = ["Sıra", "VIN No", "Durum", "EOL", "Customer", "Adet", "Tarih"]
            for col_num, header in enumerate(vin_headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = header
                cell.font = Font(size=9, bold=True, color="374151")
                cell.fill = PatternFill(start_color=self.COLOR_VIN_HEADER, end_color=self.COLOR_VIN_HEADER, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.thin_border
            
            row_num += 1
            
            # VIN detayları
            for vin_idx, vin_entry in enumerate(vin_entries, 1):
                ws.cell(row=row_num, column=1, value=vin_idx)
                ws.cell(row=row_num, column=2, value=vin_entry.VinNo)
                ws.cell(row=row_num, column=3, value=(vin_entry.Status or "PENDING").upper())
                ws.cell(row=row_num, column=4, value=vin_entry.EOLName or "-")
                ws.cell(row=row_num, column=5, value=vin_entry.CustomerReferans or "-")
                ws.cell(row=row_num, column=6, value=getattr(vin_entry, 'Adet', None) or "1")
                
                # Tarih
                if hasattr(vin_entry, 'CreatedAt') and vin_entry.CreatedAt:
                    ws.cell(row=row_num, column=7, value=vin_entry.CreatedAt.strftime('%d.%m.%Y %H:%M'))
                else:
                    ws.cell(row=row_num, column=7, value="-")
                
                # VIN detay stilleri
                for col in range(1, 8):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = PatternFill(start_color=self.COLOR_VIN_BG, end_color=self.COLOR_VIN_BG, fill_type="solid")
                    cell.border = self.thin_border
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="center" if col in [1, 3, 6] else "left", vertical="center")
                    if col == 2:  # VIN numarası
                        cell.font = Font(size=9, name="Courier New")
                
                row_num += 1
            
            # Dolly grupları arası boşluk
            row_num += 1
            dolly_index += 1
    
    def _adjust_column_widths(self, ws):
        """Sütun genişliklerini ayarla"""
        column_widths = {
            'A': 8,   # #
            'B': 25,  # DOLLY ORDER NO
            'C': 12,  # VIN SAYISI
            'D': 12,  # DURUM
            'E': 18,  # EOL
            'F': 20,  # CUSTOMER
            'G': 25   # VIN KIRILIMI
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _add_vin_by_order_sheet(self, wb, dolly_groups: Dict[str, List[Any]]):
        """
        Dolly Order No'ya göre VIN listesi tablosu ekle (basit, görselliksiz)
        Her VIN ayrı satırda - Okutma sırasına göre sıralı (CreatedAt)
        """
        # Yeni sheet oluştur
        ws = wb.create_sheet(title="VIN Listesi")
        
        # Basit başlık (1. satır)
        ws['A1'] = "DOLLY ORDER NO"
        ws['B1'] = "VIN NO"
        
        # Başlık için kalın yazı
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Tüm VIN'leri tek liste haline getir ve okutma sırasına göre sırala
        all_vins = []
        for dolly_no, vin_entries in dolly_groups.items():
            dolly_order_no = vin_entries[0].DollyOrderNo or dolly_no
            for vin_entry in vin_entries:
                all_vins.append({
                    'dolly_order_no': dolly_order_no,
                    'vin_no': vin_entry.VinNo,
                    'created_at': vin_entry.CreatedAt if hasattr(vin_entry, 'CreatedAt') else None
                })
        
        # CreatedAt'e göre sırala (okutma sırası)
        all_vins.sort(key=lambda x: x['created_at'] if x['created_at'] else datetime.min)
        
        # Sıralı VIN'leri ekle
        row_num = 2
        for vin_data in all_vins:
            ws.cell(row=row_num, column=1, value=vin_data['dolly_order_no'])
            ws.cell(row=row_num, column=2, value=vin_data['vin_no'])
            row_num += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
    
    def generate_filename(self, part_number: str) -> str:
        """Excel dosya adı oluştur"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # Part number'dan güvenli dosya adı oluştur
        safe_part = part_number.replace('/', '_').replace('\\', '_')[:50]
        return f"Dolly_Gorev_{safe_part}_{timestamp}.xlsx"

    # ---- SeferDollyEOL export (history) ----
    def export_sefer_history(self, part_number: str, records: List[Any]) -> io.BytesIO:
        """Sefer geçmişini (SeferDollyEOL) iki tablolu Excel olarak export eder."""
        if not records:
            raise ValueError("Export edilecek kayıt bulunamadı")
        
        wb = Workbook()
        ws_detail = wb.active
        ws_detail.title = "Kayıtlar"
        
        headers = [
            "Sefer No", "Plaka", "Dolly", "VIN", "Customer", "Part No", "Adet",
            "EOL", "EOL ID", "EOL Tarih", "Terminal User", "Terminal Tarih",
            "Veri Giriş", "ASN Tarih", "İrsaliye Tarih"
        ]
        self._write_header_row(ws_detail, headers)
        
        for row_idx, rec in enumerate(records, start=2):
            ws_detail.cell(row=row_idx, column=1, value=getattr(rec, "SeferNumarasi", None))
            ws_detail.cell(row=row_idx, column=2, value=getattr(rec, "PlakaNo", None))
            ws_detail.cell(row=row_idx, column=3, value=getattr(rec, "DollyNo", None))
            ws_detail.cell(row=row_idx, column=4, value=getattr(rec, "VinNo", None))
            ws_detail.cell(row=row_idx, column=5, value=getattr(rec, "CustomerReferans", None))
            ws_detail.cell(row=row_idx, column=6, value=getattr(rec, "PartNumber", None))
            ws_detail.cell(row=row_idx, column=7, value=getattr(rec, "Adet", None))
            ws_detail.cell(row=row_idx, column=8, value=getattr(rec, "EOLName", None))
            ws_detail.cell(row=row_idx, column=9, value=getattr(rec, "EOLID", None))
            ws_detail.cell(
                row=row_idx, column=10,
                value=self._fmt_dt(getattr(rec, "EOLDate", None))
            )
            ws_detail.cell(row=row_idx, column=11, value=getattr(rec, "TerminalUser", None))
            ws_detail.cell(
                row=row_idx, column=12,
                value=self._fmt_dt(getattr(rec, "TerminalDate", None))
            )
            ws_detail.cell(row=row_idx, column=13, value=getattr(rec, "VeriGirisUser", None))
            ws_detail.cell(
                row=row_idx, column=14,
                value=self._fmt_dt(getattr(rec, "ASNDate", None))
            )
            ws_detail.cell(
                row=row_idx, column=15,
                value=self._fmt_dt(getattr(rec, "IrsaliyeDate", None))
            )
        
        self._autosize(ws_detail)
        
        # Sefer bazlı özet sayfası
        ws_summary = wb.create_sheet(title="Sefer Özeti")
        summary_headers = [
            "Sefer No", "Plaka", "Part No", "Kayıt Sayısı",
            "İlk Terminal Tarihi", "Son Terminal Tarihi"
        ]
        self._write_header_row(ws_summary, summary_headers)
        
        grouped = defaultdict(list)
        for rec in records:
            key = getattr(rec, "SeferNumarasi", None) or "Bilinmiyor"
            grouped[key].append(rec)
        
        row_idx = 2
        for sefer_no, items in grouped.items():
            plaka = getattr(items[0], "PlakaNo", None)
            part = getattr(items[0], "PartNumber", None)
            terminals = [getattr(it, "TerminalDate", None) for it in items if getattr(it, "TerminalDate", None)]
            ws_summary.cell(row=row_idx, column=1, value=sefer_no)
            ws_summary.cell(row=row_idx, column=2, value=plaka)
            ws_summary.cell(row=row_idx, column=3, value=part)
            ws_summary.cell(row=row_idx, column=4, value=len(items))
            ws_summary.cell(row=row_idx, column=5, value=self._fmt_dt(min(terminals)) if terminals else None)
            ws_summary.cell(row=row_idx, column=6, value=self._fmt_dt(max(terminals)) if terminals else None)
            row_idx += 1
        
        self._autosize(ws_summary)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_sefer_filename(self, part_number: str) -> str:
        """SeferDollyEOL export dosya adı."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_part = (part_number or "PART").replace('/', '_').replace('\\', '_')[:50]
        return f"Sefer_Gecmisi_{safe_part}_{timestamp}.xlsx"

    # --- helpers ---
    def _write_header_row(self, ws, headers: List[str]):
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = Font(size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border

    def _autosize(self, ws):
        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    cell_value = cell.value or ""
                    max_length = max(max_length, len(str(cell_value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    def _fmt_dt(self, dt):
        if not dt:
            return None
        if isinstance(dt, str):
            return dt
        return dt.strftime('%d.%m.%Y %H:%M')
